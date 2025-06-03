import app_run
from django.conf import settings
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import action
from haversine import haversine, Unit
from django.db.models import Sum, Max, Min, Avg
from openpyxl import load_workbook, Workbook
from datetime import datetime
from django.db.models import Count, Q
from .services import Point, DistanceCalculator

from .models import Run, AthleteInfo, Challenge, Position, CollectibleItem, Subscriber
from .serializers import (AthleteSerializer, RunSerializer, UserSerializer, ChallengeSerializer, PositionSerializer, 
                            CollectibleItemSerializer, UserItemsSerializer, SubscribeSerializer, UserCoachSerializer,
                            UserAthleteSerializer)


@api_view(["GET"])
def company_details(request):
    return Response(
        {
            "company_name": settings.COMPANY_NAME,
            "slogan": settings.SLOGAN,
            "contacts": settings.CONTACTS,
        }
    )


class ConfigPagination(PageNumberPagination):
    page_size_query_param = "size"


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related("athlete").all()
    serializer_class = RunSerializer
    pagination_class = ConfigPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["status", "athlete"]
    ordering_fields = ["created_at"]
    ordering = ["id"]


class RunStart(APIView):

    def post(self, request, id):
        run_ = get_object_or_404(Run, id=id)
        if run_.status == "init":
            run_.status = "in_progress"
            run_.save()
        else:
            return Response(
                {"detail": "Забег не может быть начат"}, status.HTTP_400_BAD_REQUEST
            )
        return Response({"detail": "Забег успешно начат"}, status=status.HTTP_200_OK)


class RunStop(APIView):

    def post(self, request, id):
        run_ = get_object_or_404(Run, id=id)
        if run_.status == "in_progress":
            run_.status = "finished"
            start_time = Position.objects.filter(run=run_).aggregate(Min("date_time"))
            end_time = Position.objects.filter(run=run_).aggregate(Max("date_time"))
            speed_avg = Position.objects.filter(run=run_).aggregate(Avg("speed"))
            if start_time['date_time__min'] and end_time['date_time__max']:
                total_seconds = (end_time['date_time__max']-start_time['date_time__min']).total_seconds()
                run_.run_time_seconds=total_seconds
            dist=0
            for n,position in enumerate(Position.objects.filter(run=run_)):
                if n==0:
                    start_point=(position.latitude,position.longitude)
                else:
                    next_point=(position.latitude,position.longitude)
                    dist+=haversine(start_point, next_point)
                    start_point = next_point
            run_.distance=round(dist,3)
            run_.speed=round(speed_avg['speed__avg'],2)
            run_.save()
        else:
            return Response(
                {"detail": "Забег не может быть завершен"}, status.HTTP_400_BAD_REQUEST
            )
        athlete=run_.athlete
        if Run.objects.filter(status="finished", athlete=athlete).count() == 10:
            challenge = Challenge.objects.create(full_name="Сделай 10 Забегов!",athlete=athlete)
            challenge.save()
        distance = Run.objects.filter(status="finished", athlete=athlete).aggregate(Sum('distance'))
        if distance['distance__sum'] >= 50.0:
            challenge = Challenge.objects.create(full_name="Пробеги 50 километров!",athlete=athlete)
            challenge.save()           
        if run_.distance >= 2 and run_.run_time_seconds <= 600: 
            challenge = Challenge.objects.create(full_name="2 километра за 10 минут!",athlete=athlete)
            challenge.save()   
        return Response({"detail": "Забег успешно завершен"}, status=status.HTTP_200_OK)


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.annotate(runs_finished=Count('run', filter=Q(run__status='finished')))
    serializer_class = UserSerializer
    pagination_class = ConfigPagination
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["last_name", "first_name"]
    ordering_fields = ["date_joined"]
    ordering = ["id"]

    def get_queryset(self):
        qs = self.queryset
        type = self.request.query_params.get("type", None)
        if type:
            if type == "coach":
                qs = qs.filter(is_staff=True, is_superuser=False)
            else:
                qs = qs.filter(is_staff=False, is_superuser=False)
        else:
            qs = qs.filter(is_superuser=False)
        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return UserSerializer
        elif self.action == 'retrieve':
            user_ = User.objects.get(id=self.kwargs.get('pk'))
            if user_.is_staff:
                return UserCoachSerializer
            else:
                return UserAthleteSerializer
        return super().get_serializer_class()
    
class Athlete(APIView):

    def get(self, request, user_id):
        user_ = get_object_or_404(User, id=user_id)
        athlete, created = AthleteInfo.objects.get_or_create(user_id=user_)
        serializer = AthleteSerializer(athlete)
        return Response(serializer.data)

    def put(self, request, user_id):
        user_ = get_object_or_404(User, id=user_id)
        weight=request.data.get('weight')
        if weight.isdigit():
            weight=int(weight)
        else:
            return Response(
                {"detail": "Вес должен быть числом"}, status.HTTP_400_BAD_REQUEST)
        if 0<weight<900:
            athlete, created = AthleteInfo.objects.update_or_create(
                user_id=user_,
                defaults={
                    'weight': weight,
                    'goals': request.data.get('goals'),
                }
            )
        else:
            return Response(
                {"detail": "Вес введен неправильно"}, status.HTTP_400_BAD_REQUEST)
        return Response({"detail": "Операция завершилась успешно"}, status=status.HTTP_201_CREATED)


class ChallengeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer

    def get_queryset(self):
        qs = self.queryset
        athlete = self.request.query_params.get("athlete", None)
        if athlete:
            user_ = get_object_or_404(User, id=athlete)
            qs = qs.filter(athlete=user_)
        return qs


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    
    def create(self, request, *args, **kwargs):
        run_id = request.data.get("run", None)
        latitude = float(request.data.get("latitude", None))
        longitude = float(request.data.get("longitude", None))
        date_time = datetime.strptime(request.data.get("date_time", None),"%Y-%m-%dT%H:%M:%S.%f")
        if run_id and latitude and longitude:
            user_ = Run.objects.get(id=run_id).athlete
            start_point=(latitude,longitude)
            for item in CollectibleItem.objects.all():
                item_point=(item.latitude,item.longitude)
                if abs(haversine(start_point, item_point, unit=Unit.METERS)) < 100.0:
                    item.items.add(user_)
                    item.save()
            current_point=Point(latitude,longitude,date_time.replace(tzinfo=None))
            previous_data = Position.objects.filter(run=run_id).order_by("-date_time").first()
            data = request.data.copy()
            if previous_data is None:
                data['distance'] = 0.0
                data['speed'] = 0.0
            else:
                previous_point = Point(previous_data.latitude,previous_data.longitude,previous_data.date_time.replace(tzinfo=None))
                moment_data = DistanceCalculator.distance(current_point,previous_point)
                data['distance'] = previous_data.distance + moment_data['distance']
                data['speed'] = moment_data['speed']
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    def get_queryset(self):
        qs = self.queryset
        run_id = self.request.query_params.get("run", None)
        if run_id:
            try:
                run_ = Run.objects.get(id=run_id)
            except Run.DoesNotExist:
                return []
            qs = qs.filter(run=run_)
        return qs


@api_view(['POST'])
def upload_file(request):
    if "file" not in request.FILES:
        return Response({"detail": "Файл не был загружен"}, status.HTTP_400_BAD_REQUEST)
    if request.FILES["file"].content_type != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return Response({"detail": "Неверный формат файла"}, status.HTTP_400_BAD_REQUEST)
    error_list = []
    wb = load_workbook(request.FILES["file"])
    sheet = wb.active
    is_first_row = True
    for row in sheet.iter_rows(values_only=True):
        if is_first_row:
            is_first_row = False
            continue
        data_row = {'name':row[0],
                    'uid':row[1],
                    'value':row[2],
                    'latitude':row[3],
                    'longitude':row[4],
                    'picture':row[5],
        }
        serializer = CollectibleItemSerializer(data=data_row)
        if serializer.is_valid():
            serializer.save()
        else:
            error_list.append(row)
    return Response(error_list, status=status.HTTP_200_OK)


class CollectibleItemViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer

    
class SubscribeViewSet(APIView):

    def post(self, request, coach_id):
        user_ = get_object_or_404(User, id=coach_id)
        if not user_.is_staff:
            return Response({"detail": "Нельзя подписаться на атлета - только на тренера"}, status.HTTP_400_BAD_REQUEST)
        data_row = {'coach':coach_id,
                    'athlete':int(request.data['athlete']),
        }
        serializer = SubscribeSerializer(data=data_row)
        if serializer.is_valid():
            serializer.save()
            return Response({"detail": "Подписка завершилась успешно"}, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
